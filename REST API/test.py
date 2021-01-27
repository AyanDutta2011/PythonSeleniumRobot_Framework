from selenium import webdriver

driver = webdriver.Chrome(executable_path="C:\Ayan\chromedriver.exe")
driver.get("http://www.google.com/")

driver.current_url      #Gets the URL of the current page
driver.maximize_window()
driver.minimize_window()

driver.title           #title of the webpage
assert "Title" in driver.title      #to verify the title is correct or not
#Assert it is used to verify the result. If the test case fails then it will stop the execution of the test case there itself and move the control to another test case.
#Verify it is also used to verify the result. If the test case fails then it will not stop the execution of that test case.

driver.back()       #navigational command
driver.forward()    #navigational command
driver.refresh()    #navigational command

import time
time.sleep(2)
driver.implicitly_wait(10)

driver.set_script_timeout(2)        #Set the amount of time that the script should wait during an execute_async_script call before throwing an error
driver.set_page_load_timeout(10)     #Set the amount of time to wait for a page load to complete before throwing an error

ele = driver.find_element_by_name("Ayan")
ele.is_displayed()
ele.is_enabled()
ele.is_selected()      #radio button/ check box selection

from selenium.webdriver.common.by import By
driver.find_element(By.ID,"Dutta").send_keys("My Surname")      #input boxes
driver.find_element(By.ID,"Dutta").click()

from selenium.webdriver.support import expected_conditions as EC    #Explicit wait
from selenium.webdriver.support.ui import WebDriverWait
wait = WebDriverWait(driver,10)
wait.until(EC.element_to_be_clickable((By.xpath, "test")))

#drop down
from selenium.webdriver.support.ui import Select
drp = Select(ele)
drp.select_by_index(1)
drp.select_by_value("2011")
drp.select_by_visible_text("Aparna")

len(drp.options) #count no of options in drop down

from selenium.webdriver.common.by import By
links = driver.find_element(By.TAG_NAME, "a")  #no of links in a Webpage
print(len(links))

driver.find_element(By.LINK_TEXT,"Register").click()    #clicking on the link
driver.find_element(By.PARTIAL_LINK_TEXT,"Reg").click()     #clicking on the link

driver.switch_to_alert().accept()   #accept the alert
driver.switch_to_alert().dismiss()  #dismiss the alert

#frames/iframes
driver.switch_to_frame("demoFrame")
driver.find_element_by_link_text("demo").click()
driver.switch_to_default_content()
driver.switch_to_frame("demoFrame2")
driver.find_element_by_link_text("demo2").click()

driver.current_window_handle    #select current window
exp = driver.window_handles     #return all the values of open browsers
for i in exp:
    driver.switch_to_window(i)  #switch to a open window
    print(driver.title)
    if driver.title == "Ayan":
        driver.quit()           #close present browser

driver.find_element_by_id("test").text         #get text of web element

driver.execute_script("window.scrollBy(0,1000)","") #scroll down by pixel
driver.execute_script("arguments[0].scrollIntoView();", ele)    #scroll down till element view
driver.execute_script("window.scrollBy(0,document.body.scrollHeight)")  #page end scroll down

ele1 = driver.find_element_by_name("Ayan1")
ele2 = driver.find_element_by_name("Ayan2")
ele3 = driver.find_element_by_name("Ayan3")

# Mouse actions
from selenium.webdriver import ActionChains
action = ActionChains(driver)
action.move_to_element(ele1).move_to_element(ele2).move_to_element(ele3).click().perform()  #Mouse hover
action.double_click(ele3).perform()     #double click
action.context_click(ele2).perform()    #right click
action.drag_and_drop(ele2,ele3)         #drag and drop (source to target)
action.move_by_offset()  #move the mouse control to specific cooridinate in selenium

driver.find_element_by_id("test").send_keys("C:/Users/Ayan Dutta/Pictures/Screenshots/test.jpg")   #upload files

#Download in Chrome
from selenium.webdriver.chrome.options import Options
chromeOptions = Options()
chromeOptions.add_experimental_option("prefs", ("download.default_directory": "C:\Download"))

#cookies
driver.get_cookies()
add = {'name':'Yamaha', 'value':'12345'}
driver.add_cookie(add)
driver.delete_cookie("Yamaha")

#screenshot
driver.save_screenshot("C:\demo\page.png")
driver.save_screenshot_as_file("C:\demo\page2.png")

driver.quit()   #to close current browser
driver.close()  #to close all open browser

#Working with HTML table
#tr = table row
#td = table data
#th = table header cols

rows = len(driver.find_elements_by_xpath("/html/body/table/tbody/tr"))       #all the rows
cols = len(driver.find_elements_by_xpath("/html/body/table/tbody/tr[1]/th"))        #all the cols
/html/body/table/tbody/tr[1]    #first row selected
/html/body/table/tbody/tr[1]/th[1]      #first table header selected
for r in range(1,rows+1):
    for c in range(1,cols+1):
        value = driver.find_element_by_xpath("/html/body/table/tbody/tr["+str(r)+"]/th["+str(c)+"] ").text
        print(value, end= '   ')

#How to create xpath:
//input[@name='q']     #create xpath of a textbox   //htmltag[attribute]
//input[@type='submit' and @value='Login']       #create xpath  //htmltag[attribute]
//a[text()='Features']          #create xpath of a link
//button[@type='button' and @class='btn']         #create xpath of a button
//button[contains(text(),'Sign Up')]         #create xpath of a button using contains

//div[@class='test']//button[@type='button' and @class='btn']   #complex xpath create

#goto a web table and click on the checkbox

//a[text()='Features']//parent::td[@class='test1']//preceding-sibling::td[@class='test2']//input[@name='test_id']

# Absolute Xpath: It contains the complete path from the Root Element to the desire element. it starts with single /
# Relative Xpath: This is more like starting simply by referencing the element you want and go from the particular location. it starts with single //
# You use always the Relative Path for testing of an element.

#To achieve this type of issue we can pass username and pwd with URL
http://username:password@url

###############################################################

import openpyxl

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)
def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)
def readData(file,sheetName,rownum,columno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columno).value
def writeData(file,sheetName,rownum,columno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columno).value = data
    workbook.save(file)

read = ref.readData(path,'Connectivity OBS List',r,3)
write = ref.writeData(path, 'Connectivity OBS List', r, 11, date)

